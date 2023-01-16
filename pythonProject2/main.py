import openpyxl
from openpyxl.styles import Font, PatternFill

year = input("Input year:\n")
wb = openpyxl.load_workbook("Паспорта.xlsx")
sheet = wb["Лист1"]
maxRows = sheet.max_row
needMonth = ["01", "02", "03"]
cnt = 0
for i in range(2, maxRows + 1):
    cellValue = sheet.cell(row = i, column = 7)
    date = str(cellValue.value).split('-')
    if (date[0] == year or (date[0] == str(int(year) + 1) and date[1] in needMonth)):
        cellValue.fill = PatternFill(start_color = "ffff00", end_color = "ffff00", fill_type = "solid")
    cnt += 1
    print(f"{cnt} of {maxRows}")
print("Done")
sheet.cell(row = maxRows, column = 11).fill = PatternFill(start_color = "ff0000", end_color = "ff0000", fill_type = "solid")
wb.save("Паспорта.xlsx")